'''
zipファイルを解凍してできたエクセルから昨日分の30分値を取得する
そして、その情報を今月分の30分値をまとめたエクセルに転記する
'''
from config import Config
import pandas as pd
from openpyxl import load_workbook

class ExcelWriting:
    def __init__(self, excel_file, output_excel_file, filtering):
        self.excel_file = excel_file
        self.output_excel_file = output_excel_file
        self.filtering = filtering
        self.config = Config()
        self.yesterday_tsuki_nichi = self.config.yesterday_tsuki_nichi
        
    def excel_reading(self):
        #スクレイピングで取得したエクセルから昨日分の30分値を取得する
        df_raw = pd.read_excel(self.excel_file, header=None)
        df_raw.columns = df_raw.iloc[1]
        df = df_raw.iloc[2:].reset_index(drop=True)

        self.df_get = df.loc[
            df['/JPMGRP/JPTRM/JPM00010/JPMR00010/JPM00011/JPMR00011/JP06120'] == self.filtering
        ]['/JPMGRP/JPTRM/JPM00010/JPMR00010/JPM00011/JPMR00011/JP06123'].astype(float).tolist()  
    
    def excel_copy(self):
        #今月分のエクセルに昨日分の30分値を転記する
        wb = load_workbook(self.output_excel_file)
        ws = wb.active

        # 日付列を探す（1行目のセルに yesterday_tsuki_nichi と一致する列）
        date_col = None
        for col in range(2, ws.max_column + 1):  # A列は時間なのでB列以降
            if ws.cell(row=1, column=col).value == self.yesterday_tsuki_nichi:
                date_col = col
                break

        if date_col is None:
            raise ValueError(f"列「{self.yesterday_tsuki_nichi}」が見つかりませんでした。")

        # データ入力（0〜47行目）
        for i in range(48):
            ws.cell(row=i+2, column=date_col, value=self.df_get[i] if i < len(self.df_get) else 0)

        # 列合計（48行目 = 49行目、Excel的には row=50）
        total = sum(self.df_get)
        ws.cell(row=50, column=date_col, value=total)

        # 各行の合計を "ALL" 列に書き込む
        # "ALL" 列を見つける
        all_col = None
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "ALL":
                all_col = col
                break

        if all_col is not None:
            for row in range(2, 51):  # 1行目はヘッダーなので除く
                row_sum = 0
                for col in range(2, all_col):  # "ALL"列を含まない範囲
                    val = ws.cell(row=row, column=col).value
                    row_sum += float(val) if val not in (None, "") else 0
                ws.cell(row=row, column=all_col, value=row_sum)

        # 保存
        wb.save(self.output_excel_file)
        
    def excel_writing(self):
        #この関数で実行する
        self.excel_reading()
        self.excel_copy()