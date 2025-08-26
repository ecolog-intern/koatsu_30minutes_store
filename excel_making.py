#月ごとのエクセルファイルを作成する
from datetime import datetime, timedelta
import calendar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class MonthExcelMaking:
    def __init__(self, full_path):
        self.full_path = full_path
        self.yesterday = datetime.today() - timedelta(days=1)
        self.year, self.month = self.yesterday.year, self.yesterday.month
        _, self.last_day = calendar.monthrange(self.year, self.month)
        
        self.date_list = [f"{self.month}月{day}日" for day in range(1, self.last_day + 1)] + ['ALL']
        self.time_slots = [
            f"{(datetime.strptime('00:00', '%H:%M') + timedelta(minutes=30*i)).strftime('%H:%M')}-"
            f"{(datetime.strptime('00:00', '%H:%M') + timedelta(minutes=30*(i+1))).strftime('%H:%M')}"
            for i in range(48)
        ] + ['ALL']
        
    def excel_create(self):
        #full_pathにエクセルを作成する
        df = pd.DataFrame(index=self.time_slots, columns=self.date_list)
        df.to_excel(self.full_path)
        
    def excel_styles(self):
        #エクセルをopenpyxlで読み込んでデザインを整える
        wb = load_workbook(self.full_path)
        ws = wb.active
        ws["A1"] = "時間"

        # スタイル定義
        base_font = Font(name="BIZ UDPゴシック", size=8)
        bold_font = Font(name="BIZ UDPゴシック", size=8, bold=True)
        header_fill = PatternFill(start_color="EAF5EA", end_color="EAF5EA", fill_type="solid")
        all_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        no_fill = PatternFill(fill_type=None, start_color="FFFFFF", end_color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # フォント適用（全体）
        for row in ws.iter_rows():
            for cell in row:
                cell.font = base_font

        # 1行目（ヘッダー）：塗りつぶし、太字、外枠
        for cell in ws[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = bold_font
                cell.border = thin_border

        # A列 2〜49：時間帯列 → 塗りつぶしなし、太字、外枠
        for row in ws.iter_rows(min_row=2, max_row=49, min_col=1, max_col=1):
            for cell in row:
                cell.font = bold_font
                cell.fill = no_fill  # 完全に塗りつぶしなし
                cell.border = thin_border

        # A列 50行目（ALL）：グレー、太字、外枠
        cell = ws["A50"]
        cell.font = bold_font
        cell.fill = all_fill
        cell.border = thin_border

        # "ALL"を含む行：グレー、太字、外枠
        for row in ws.iter_rows():
            if any(cell.value == "ALL" for cell in row):
                for cell in row:
                    cell.font = bold_font
                    cell.fill = all_fill
                    cell.border = thin_border

        # "ALL"を含む列：グレー、太字、外枠
        for col in ws.iter_cols():
            if any(cell.value == "ALL" for cell in col):
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.font = bold_font
                    cell.fill = all_fill
                    cell.border = thin_border
                ws.column_dimensions[col_letter].width = 14

        # 列幅自動調整（最小12）
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

        # 保存
        wb.save(self.full_path)
    
    def excel_making(self):
        #外部からの呼び出し用
        self.excel_create()
        self.excel_styles()
        